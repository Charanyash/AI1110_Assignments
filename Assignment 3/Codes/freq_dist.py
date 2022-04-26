import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from collections import Counter

df=pd.read_excel('raw.xlsx','Sheet1',header=None)

a = df.to_numpy()
df=pd.DataFrame(a)
newlst = []
df.applymap(lambda x: newlst.append(x))
newlst.sort()

cou=Counter(newlst)
marks=list(cou.keys())
freq=list(cou.values())

total = len(newlst)

df1 = pd.DataFrame()
df1['marks']=marks
df1['Number of students(frequency)']=freq
df1.to_excel('Freq_dist.xlsx',index=False)

wb=openpyxl.load_workbook("Freq_dist.xlsx")
ws=wb['Sheet1']
ws['A14']='Total'
ws['B14']=total
ws.column_dimensions['B'].width = 25
ws.row_dimensions[1].height = 30
ws['B1'].alignment = Alignment(wrap_text=True)

wb.save("Freq_dist.xlsx")
wb.close()

